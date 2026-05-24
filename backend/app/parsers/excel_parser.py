"""ExcelParser - 基于 openpyxl，纯 Python 兼容 LoongArch.

解析 .xlsx 文件，提取所有工作表的表格数据。
"""

from __future__ import annotations

import io

from app.parsers.base import ParsedDocument, ParserError


class ExcelParser:
    file_type = "xlsx"

    async def parse(self, content: bytes) -> ParsedDocument:
        try:
            from openpyxl import load_workbook
        except ImportError as e:
            raise ParserError("openpyxl 未安装", field="dependency") from e

        try:
            wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        except Exception as e:  # noqa: BLE001
            raise ParserError(f"xlsx 解析失败: {e}", field="file") from e

        tables: list[list[list[str]]] = []
        raw_text_parts: list[str] = []
        paragraphs: list[str] = []

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            paragraphs.append(f"[工作表: {sheet_name}]")
            raw_text_parts.append(f"=== 工作表: {sheet_name} ===")

            sheet_rows: list[list[str]] = []
            for row in ws.iter_rows(values_only=True):
                cells = [str(c) if c is not None else "" for c in row]
                # 跳过全空行
                if not any(cells):
                    continue
                sheet_rows.append(cells)
                row_text = " | ".join(c for c in cells if c)
                if row_text:
                    raw_text_parts.append(row_text)

            if sheet_rows:
                tables.append(sheet_rows)

        wb.close()

        return ParsedDocument(
            raw_text="\n".join(raw_text_parts),
            paragraphs=paragraphs,
            tables=tables,
            page_count=len(wb.sheetnames) if wb.sheetnames else 0,
        )
