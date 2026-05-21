"""ImportService - Epic 25.2 / 25.3 / 25.5."""

from __future__ import annotations

import io
from datetime import UTC, datetime
from typing import Any

from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.core.exceptions import (
    AuthorizationError,
    BusinessRuleError,
    ResourceNotFoundError,
)
from app.core.logging import get_logger
from app.core.security import hash_password
from app.models.course import Class, ClassMembership
from app.models.import_job import ImportJob, ImportRecord
from app.models.user import User


log = get_logger(__name__)

IMPORT_LIMIT = 500
IMPORT_LIMIT_ERROR_CODE = "IMPORT_LIMIT_EXCEEDED"


class ImportLimitExceededError(BusinessRuleError):
    error_code = IMPORT_LIMIT_ERROR_CODE


class ImportService:
    # ============== 25.2 用户批量导入 ==============

    async def start_user_import(
        self,
        db: AsyncSession,
        *,
        operator: User,
        rows: list[dict[str, Any]],
    ) -> ImportJob:
        if len(rows) > IMPORT_LIMIT:
            raise ImportLimitExceededError(
                f"导入行数 {len(rows)} 超过上限 {IMPORT_LIMIT}",
                field="rows",
            )
        job = ImportJob(
            operator_id=operator.id,
            job_type="user",
            status="processing",
            total_count=len(rows),
        )
        db.add(job)
        await db.flush()

        success = 0
        failed = 0
        for idx, row in enumerate(rows, start=1):
            sp = await db.begin_nested()
            try:
                username = str(row.get("username", "")).strip()
                role = str(row.get("role", "student")).strip()
                display_name = str(row.get("display_name", username)).strip()
                password = str(row.get("password", "Pa$$w0rd2024"))
                if not username or len(username) < 3:
                    raise ValueError("username 长度不足")
                if role not in ("student", "teacher", "admin"):
                    raise ValueError(f"role {role} 不合法")
                if len(password) < 8:
                    raise ValueError("password 长度不足 8")
                # 检查重复
                existing = (
                    await db.execute(
                        select(User).where(User.username == username)
                    )
                ).scalar_one_or_none()
                if existing is not None:
                    raise ValueError(f"用户名 {username} 已存在")

                u = User(
                    username=username,
                    display_name=display_name,
                    password_hash=hash_password(password),
                    role=role,
                    is_active=True,
                )
                db.add(u)
                await db.flush()
                await sp.commit()
                success += 1
                db.add(
                    ImportRecord(
                        job_id=job.id,
                        row_number=idx,
                        status="success",
                    )
                )
            except Exception as e:  # noqa: BLE001
                await sp.rollback()
                log.warning("import.user.row_failed", row=idx, error=str(e), error_type=type(e).__name__)
                db.add(
                    ImportRecord(
                        job_id=job.id,
                        row_number=idx,
                        status="failed",
                        error_message=str(e),
                    )
                )
                failed += 1

        job.success_count = success
        job.failed_count = failed
        job.status = "done"
        job.completed_at = datetime.now(UTC)
        await db.flush()
        log.info(
            "import.user.done",
            job_id=job.id,
            success=success,
            failed=failed,
        )
        return job

    # ============== 25.3 学生班级导入 ==============

    async def start_class_student_import(
        self,
        db: AsyncSession,
        *,
        operator: User,
        class_id: int,
        usernames: list[str],
    ) -> ImportJob:
        if len(usernames) > IMPORT_LIMIT:
            raise ImportLimitExceededError(
                f"导入行数 {len(usernames)} 超过上限"
            )
        cls = await db.get(Class, class_id)
        if cls is None:
            raise ResourceNotFoundError(f"class {class_id} not found")
        if operator.role != "admin" and cls.teacher_id != operator.id:
            raise AuthorizationError("无权操作其他教师班级")

        job = ImportJob(
            operator_id=operator.id,
            job_type="student_to_class",
            status="processing",
            total_count=len(usernames),
        )
        db.add(job)
        await db.flush()

        success = 0
        failed = 0
        for idx, username in enumerate(usernames, start=1):
            try:
                u = (
                    await db.execute(
                        select(User).where(User.username == username)
                    )
                ).scalar_one_or_none()
                if u is None or u.role != "student":
                    raise ValueError(f"username {username} 不存在或非 student")
                # 已在班级 → 跳过且计成功
                exists = (
                    await db.execute(
                        select(ClassMembership).where(
                            ClassMembership.class_id == class_id,
                            ClassMembership.student_id == u.id,
                        )
                    )
                ).scalar_one_or_none()
                if exists is None:
                    db.add(
                        ClassMembership(
                            class_id=class_id, student_id=u.id
                        )
                    )
                db.add(
                    ImportRecord(
                        job_id=job.id, row_number=idx, status="success"
                    )
                )
                success += 1
            except Exception as e:  # noqa: BLE001
                db.add(
                    ImportRecord(
                        job_id=job.id,
                        row_number=idx,
                        status="failed",
                        error_message=str(e),
                    )
                )
                failed += 1

        job.success_count = success
        job.failed_count = failed
        job.status = "done"
        job.completed_at = datetime.now(UTC)
        await db.flush()
        return job


# ============== 25.4 模板生成 ==============


def make_user_template_xlsx() -> bytes:
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.title = "users"
    ws.append(["username", "display_name", "role", "password"])
    ws.append(["zhangsan", "张三", "student", "Pa$$w0rd2024"])
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def make_student_template_xlsx() -> bytes:
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.title = "students"
    ws.append(["username"])
    ws.append(["zhangsan"])
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ============== 25.5 导出 ==============


async def export_class_students_xlsx(
    db: AsyncSession, *, class_id: int, operator: User
) -> bytes:
    cls = await db.get(Class, class_id)
    if cls is None:
        raise ResourceNotFoundError(f"class {class_id} not found")
    if operator.role != "admin" and cls.teacher_id != operator.id:
        raise AuthorizationError("无权导出其他教师班级")

    rows = list(
        (
            await db.execute(
                select(ClassMembership, User)
                .join(User, User.id == ClassMembership.student_id)
                .where(ClassMembership.class_id == class_id)
            )
        ).all()
    )

    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.title = "students"
    ws.append(["username", "display_name"])
    for _, u in rows:
        ws.append([u.username, u.display_name])
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def parse_user_xlsx(data: bytes) -> list[dict[str, Any]]:
    """读取用户导入 xlsx，返回 row dicts."""
    from openpyxl import load_workbook

    wb = load_workbook(filename=io.BytesIO(data))
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [str(h) if h is not None else "" for h in rows[0]]
    out: list[dict[str, Any]] = []
    for r in rows[1:]:
        if all(c is None for c in r):
            continue
        d = {h: (str(v) if v is not None else "") for h, v in zip(headers, r, strict=False)}
        out.append(d)
    return out


def parse_student_xlsx(data: bytes) -> list[str]:
    from openpyxl import load_workbook

    wb = load_workbook(filename=io.BytesIO(data))
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    out: list[str] = []
    for r in rows[1:]:
        if r and r[0]:
            out.append(str(r[0]).strip())
    return out


__all__ = [
    "ImportLimitExceededError",
    "ImportService",
    "export_class_students_xlsx",
    "make_student_template_xlsx",
    "make_user_template_xlsx",
    "parse_student_xlsx",
    "parse_user_xlsx",
]
