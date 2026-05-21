"""Excel 报表渲染（Epic 20.3）."""

from __future__ import annotations

import io
from typing import Any


def render_statistics_xlsx(rows: list[dict[str, Any]]) -> bytes:
    """生成统计 xlsx 字节流.

    Sheet1 数据明细，行字段：
    - student_name, total_score, status
    - 任意维度评分动态列
    """
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "数据明细"

    # 收集动态维度列
    dim_set: list[str] = []
    for r in rows:
        for d in r.get("dimensions", []):
            n = d.get("name")
            if n and n not in dim_set:
                dim_set.append(n)

    headers = ["学生姓名", "综合分", "状态"] + dim_set
    ws.append(headers)
    header_font = Font(bold=True)
    header_fill = PatternFill("solid", fgColor="E2E8F0")
    for col_idx, _ in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    # 数据行
    for r in rows:
        dim_map = {d.get("name"): d for d in r.get("dimensions", [])}
        row_vals: list[Any] = [
            r.get("student_name", ""),
            r.get("total_score", "未评价"),
            r.get("status", ""),
        ]
        for n in dim_set:
            d = dim_map.get(n)
            if d is None:
                row_vals.append("未评价")
            else:
                row_vals.append(d.get("teacher_score") or d.get("ai_score") or "")
        ws.append(row_vals)

    # 冻结表头 + 自动列宽
    ws.freeze_panes = "A2"
    for col_idx in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = 16

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


__all__ = ["render_statistics_xlsx"]
