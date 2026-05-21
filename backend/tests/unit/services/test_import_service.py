"""Epic 25.1/25.2/25.3/25.5 验收：ImportService."""

from __future__ import annotations

import io
from collections.abc import AsyncIterator

import pytest
from openpyxl import Workbook, load_workbook
from sqlalchemy import text
from sqlalchemy.ext.asyncio import AsyncSession, async_sessionmaker, create_async_engine

from app.core.database import Base
from app.core.exceptions import AuthorizationError
from app.services.import_service import (
    ImportLimitExceededError,
    ImportService,
    export_class_students_xlsx,
    make_student_template_xlsx,
    make_user_template_xlsx,
    parse_student_xlsx,
    parse_user_xlsx,
)
from tests.factories.org_factory import ClassFactory
from tests.factories.user_factory import (
    AdminFactory,
    TeacherFactory,
    UserFactory,
)


pytestmark = pytest.mark.unit


@pytest.fixture()
async def session() -> AsyncIterator[AsyncSession]:
    engine = create_async_engine("sqlite+aiosqlite:///:memory:")
    async with engine.begin() as conn:
        await conn.execute(text("PRAGMA foreign_keys = ON"))
        await conn.run_sync(Base.metadata.create_all)
    SessionLocal = async_sessionmaker(engine, expire_on_commit=False, class_=AsyncSession)
    async with SessionLocal() as s:
        yield s
    await engine.dispose()


class TestUserImport:
    async def test_partial_failure(self, session: AsyncSession) -> None:
        admin = await AdminFactory.create_async(session)
        await session.commit()
        rows = [
            {"username": "user_a", "display_name": "U1", "role": "student", "password": "Pa$$w0rd2024"},
            {"username": "user_b", "display_name": "U2", "role": "student", "password": "Pa$$w0rd2024"},
            {"username": "x", "display_name": "X", "role": "student", "password": "Pa$$w0rd2024"},  # username 太短
            {"username": "user_c", "display_name": "U3", "role": "bad_role", "password": "Pa$$w0rd2024"},  # role 不合法
            {"username": "user_a", "display_name": "Dup", "role": "student", "password": "Pa$$w0rd2024"},  # 重复
        ]
        svc = ImportService()
        job = await svc.start_user_import(session, operator=admin, rows=rows)
        await session.commit()
        assert job.total_count == 5
        assert job.success_count == 2
        assert job.failed_count == 3
        assert job.status == "done"

    async def test_over_limit_raises(self, session: AsyncSession) -> None:
        admin = await AdminFactory.create_async(session)
        await session.commit()
        rows = [
            {"username": f"u{i}", "role": "student"} for i in range(501)
        ]
        svc = ImportService()
        with pytest.raises(ImportLimitExceededError):
            await svc.start_user_import(session, operator=admin, rows=rows)


class TestClassStudentImport:
    async def test_unknown_username_failed(
        self, session: AsyncSession
    ) -> None:
        teacher = await TeacherFactory.create_async(session)
        cls = await ClassFactory.create_async(session, teacher=teacher)
        await UserFactory.create_async(session, username="real")
        await session.commit()
        svc = ImportService()
        job = await svc.start_class_student_import(
            session,
            operator=teacher,
            class_id=cls.id,
            usernames=["real", "ghost"],
        )
        await session.commit()
        assert job.success_count == 1
        assert job.failed_count == 1

    async def test_already_in_class_counts_success(
        self, session: AsyncSession
    ) -> None:
        teacher = await TeacherFactory.create_async(session)
        cls = await ClassFactory.create_async(session, teacher=teacher)
        student = await UserFactory.create_async(session, username="alice")
        await session.commit()
        svc = ImportService()
        # 第一次导入
        await svc.start_class_student_import(
            session,
            operator=teacher,
            class_id=cls.id,
            usernames=["alice"],
        )
        # 重复导入仍然成功（idempotent）
        job = await svc.start_class_student_import(
            session,
            operator=teacher,
            class_id=cls.id,
            usernames=["alice"],
        )
        await session.commit()
        assert job.success_count == 1
        assert job.failed_count == 0

    async def test_other_teacher_forbidden(
        self, session: AsyncSession
    ) -> None:
        teacher_a = await TeacherFactory.create_async(session)
        teacher_b = await TeacherFactory.create_async(session)
        cls = await ClassFactory.create_async(session, teacher=teacher_a)
        await session.commit()
        svc = ImportService()
        with pytest.raises(AuthorizationError):
            await svc.start_class_student_import(
                session,
                operator=teacher_b,
                class_id=cls.id,
                usernames=[],
            )


class TestTemplates:
    def test_user_template_can_open(self) -> None:
        data = make_user_template_xlsx()
        wb = load_workbook(io.BytesIO(data))
        ws = wb.active
        assert ws.cell(1, 1).value == "username"

    def test_student_template_can_open(self) -> None:
        data = make_student_template_xlsx()
        wb = load_workbook(io.BytesIO(data))
        assert wb.active.cell(1, 1).value == "username"


class TestParse:
    def test_parse_user_xlsx_valid_rows(self) -> None:
        wb = Workbook()
        ws = wb.active
        ws.append(["username", "display_name", "role"])
        ws.append(["u1", "U1", "student"])
        ws.append([None, None, None])  # 空行
        ws.append(["u2", "U2", "teacher"])
        buf = io.BytesIO()
        wb.save(buf)
        rows = parse_user_xlsx(buf.getvalue())
        assert len(rows) == 2
        assert rows[0]["username"] == "u1"

    def test_parse_student_xlsx(self) -> None:
        wb = Workbook()
        ws = wb.active
        ws.append(["username"])
        ws.append(["u1"])
        ws.append(["u2"])
        buf = io.BytesIO()
        wb.save(buf)
        names = parse_student_xlsx(buf.getvalue())
        assert names == ["u1", "u2"]


class TestExport:
    async def test_export_class_students(
        self, session: AsyncSession
    ) -> None:
        from app.models.course import ClassMembership

        teacher = await TeacherFactory.create_async(session)
        cls = await ClassFactory.create_async(session, teacher=teacher)
        s1 = await UserFactory.create_async(session, username="s1")
        s2 = await UserFactory.create_async(session, username="s2")
        session.add_all(
            [
                ClassMembership(class_id=cls.id, student_id=s1.id),
                ClassMembership(class_id=cls.id, student_id=s2.id),
            ]
        )
        await session.commit()
        data = await export_class_students_xlsx(
            session, class_id=cls.id, operator=teacher
        )
        wb = load_workbook(io.BytesIO(data))
        ws = wb.active
        names = [ws.cell(i, 1).value for i in range(2, 4)]
        assert "s1" in names and "s2" in names

    async def test_export_other_teacher_forbidden(
        self, session: AsyncSession
    ) -> None:
        teacher_a = await TeacherFactory.create_async(session)
        teacher_b = await TeacherFactory.create_async(session)
        cls = await ClassFactory.create_async(session, teacher=teacher_a)
        await session.commit()
        with pytest.raises(AuthorizationError):
            await export_class_students_xlsx(
                session, class_id=cls.id, operator=teacher_b
            )
