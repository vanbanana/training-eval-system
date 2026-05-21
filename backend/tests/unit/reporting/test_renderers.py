"""Epic 20.1 / 20.2 / 20.3 验收：图表/PDF/Excel 渲染."""

from __future__ import annotations

import io

import pytest

from app.reporting import (
    render_bar_chart_svg,
    render_line_chart_svg,
    render_personal_html,
    render_personal_pdf,
    render_radar_chart_svg,
    render_statistics_xlsx,
)


pytestmark = pytest.mark.unit


class TestChartRenderer:
    def test_radar_returns_svg(self) -> None:
        s = render_radar_chart_svg(["代码", "文档", "测试"], [80, 70, 90])
        assert s.startswith("<svg")
        assert "代码" in s

    def test_bar_returns_svg(self) -> None:
        s = render_bar_chart_svg(["A", "B"], [50, 90])
        assert "<rect" in s

    def test_line_returns_svg(self) -> None:
        s = render_line_chart_svg([("第1次", 80), ("第2次", 90)])
        assert "<polyline" in s

    def test_radar_invalid_input(self) -> None:
        with pytest.raises(ValueError):
            render_radar_chart_svg(["a"], [1, 2])


class TestPdfRenderer:
    def test_html_includes_data(self) -> None:
        h = render_personal_html(
            {
                "student_name": "张三",
                "task_name": "任务1",
                "total_score": 88,
                "dimensions": [
                    {"name": "代码质量", "ai_score": 80, "teacher_score": 90, "rationale": "ok"}
                ],
                "teacher_comment": "不错",
            }
        )
        assert "张三" in h
        assert "代码质量" in h

    def test_pdf_returns_bytes(self) -> None:
        out = render_personal_pdf(
            {
                "student_name": "张三",
                "task_name": "任务1",
                "total_score": 88,
                "dimensions": [
                    {"name": "代码质量", "ai_score": 80, "teacher_score": 90, "rationale": "ok"}
                ],
                "teacher_comment": "不错",
            }
        )
        assert isinstance(out, bytes)
        assert len(out) > 100


class TestExcelRenderer:
    def test_xlsx_can_reload(self) -> None:
        from openpyxl import load_workbook

        rows = [
            {
                "student_name": "张三",
                "total_score": 88,
                "status": "auto_scored",
                "dimensions": [
                    {"name": "代码", "ai_score": 80},
                    {"name": "文档", "ai_score": 70},
                ],
            },
            {
                "student_name": "李四",
                "total_score": 75,
                "status": "auto_scored",
                "dimensions": [
                    {"name": "代码", "ai_score": 75},
                    {"name": "文档", "ai_score": 70},
                ],
            },
        ]
        out = render_statistics_xlsx(rows)
        wb = load_workbook(filename=io.BytesIO(out))
        ws = wb.active
        # 表头
        assert ws.cell(row=1, column=1).value == "学生姓名"
        # 数据
        names = [ws.cell(row=i, column=1).value for i in range(2, 4)]
        assert "张三" in names
        assert "李四" in names
