"""Epic 20.6 验收：报表流程集成."""

from __future__ import annotations

import io

import pytest
from openpyxl import load_workbook
from sqlalchemy.ext.asyncio import AsyncSession

from app.models.evaluation import DimensionScore, Evaluation
from app.models.upload import Upload
from app.services.report_service import ReportService
from tests.factories.task_factory import TrainingTaskFactory
from tests.factories.user_factory import TeacherFactory, UserFactory


pytestmark = pytest.mark.integration


class TestReportFlow:
    async def test_generates_for_30_students(
        self, sqlite_session: AsyncSession
    ) -> None:
        """Given 30 学生评价 When 导出 xlsx Then 行数=30+1表头."""
        teacher = await TeacherFactory.create_async(sqlite_session)
        task = await TrainingTaskFactory.create_async(
            sqlite_session, teacher=teacher, with_dimensions=2
        )
        for i in range(30):
            student = await UserFactory.create_async(sqlite_session)
            upload = Upload(
                task_id=task.id,
                student_id=student.id,
                filename="r",
                file_type="docx",
                file_size=10,
                storage_path=f"x{i}",
                parse_status="parsed",
            )
            sqlite_session.add(upload)
            await sqlite_session.flush()
            ev = Evaluation(
                task_id=task.id,
                student_id=student.id,
                upload_id=upload.id,
                status="auto_scored",
                total_score=70.0 + i % 30,
            )
            sqlite_session.add(ev)
            await sqlite_session.flush()
            for d in task.dimensions:
                sqlite_session.add(
                    DimensionScore(
                        evaluation_id=ev.id,
                        dimension_id=d.id,
                        ai_score=70.0 + i % 30,
                        rationale="ok",
                    )
                )
        await sqlite_session.commit()

        svc = ReportService()
        xlsx, _ = await svc.generate_statistics_xlsx(
            sqlite_session, task_id=task.id, operator=teacher
        )
        wb = load_workbook(filename=io.BytesIO(xlsx))
        ws = wb.active
        # 表头 1 行 + 30 学生 = 31 行
        assert ws.max_row == 31
